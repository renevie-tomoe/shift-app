"""
シフト表をExcelに出力する
"""
import calendar
import os
from datetime import date
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

WEEKDAY_JP = ["月", "火", "水", "木", "金", "土", "日"]
ROLE_ORDER = ["店長", "マネージャー", "ロングパート", "ショートパート", "アルバイト"]

COLOR_HEADER = "4F81BD"
COLOR_SAT = "DCE6F1"
COLOR_SUN = "F2DCDB"
COLOR_WORK = "FFFFFF"
COLOR_OFF = "D9D9D9"


def _thin_border():
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def export_to_excel(year: int, month: int, employees: list, shift_result: dict, output_dir: str) -> str:
    num_days = calendar.monthrange(year, month)[1]
    days = list(range(1, num_days + 1))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{month}月シフト"

    border = _thin_border()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
    sat_fill = PatternFill("solid", fgColor=COLOR_SAT)
    sun_fill = PatternFill("solid", fgColor=COLOR_SUN)
    off_fill = PatternFill("solid", fgColor=COLOR_OFF)

    # ---- ヘッダー行 ----
    ws.cell(1, 1, f"{year}年{month}月 シフト表").font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + num_days)

    # 行2: ラベル + 日付
    ws.cell(2, 1, "ロール").border = border
    ws.cell(2, 2, "名前").border = border
    ws.cell(2, 3, "勤務日数").border = border

    for i, d in enumerate(days):
        col = 4 + i
        weekday = calendar.weekday(year, month, d)
        cell_date = ws.cell(2, col, d)
        cell_date.border = border
        cell_date.alignment = Alignment(horizontal="center")
        if weekday == 5:
            cell_date.fill = sat_fill
        elif weekday == 6:
            cell_date.fill = sun_fill
        else:
            cell_date.fill = header_fill
            cell_date.font = header_font

    # 行3: 曜日
    ws.cell(3, 1, "").border = border
    ws.cell(3, 2, "").border = border
    ws.cell(3, 3, "").border = border
    for i, d in enumerate(days):
        col = 4 + i
        weekday = calendar.weekday(year, month, d)
        cell_wd = ws.cell(3, col, WEEKDAY_JP[weekday])
        cell_wd.border = border
        cell_wd.alignment = Alignment(horizontal="center")
        if weekday == 5:
            cell_wd.fill = sat_fill
        elif weekday == 6:
            cell_wd.fill = sun_fill

    # ---- データ行 ----
    sorted_employees = sorted(employees, key=lambda e: ROLE_ORDER.index(e["role"]) if e["role"] in ROLE_ORDER else 99)

    row = 4
    current_role = None
    for e in sorted_employees:
        name = e["name"]
        role = e["role"]

        if role != current_role:
            # ロール区切り行
            role_cell = ws.cell(row, 1, role)
            role_cell.font = Font(bold=True)
            role_cell.fill = PatternFill("solid", fgColor="BDD7EE")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3 + num_days)
            row += 1
            current_role = role

        ws.cell(row, 1, role).border = border
        ws.cell(row, 2, name).border = border

        work_days = 0
        for i, d in enumerate(days):
            col = 4 + i
            weekday = calendar.weekday(year, month, d)
            is_working = shift_result.get(name, {}).get(d, False)

            label = "○" if is_working else "×"
            if is_working:
                work_days += 1

            cell = ws.cell(row, col, label)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

            if not is_working:
                cell.fill = off_fill
            elif weekday == 5:
                cell.fill = sat_fill
            elif weekday == 6:
                cell.fill = sun_fill

        ws.cell(row, 3, work_days).border = border
        ws.cell(row, 3).alignment = Alignment(horizontal="center")
        row += 1

    # 列幅調整
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 8
    for i in range(num_days):
        ws.column_dimensions[get_column_letter(4 + i)].width = 3.5

    # 行高さ
    for r in range(1, row):
        ws.row_dimensions[r].height = 18

    os.makedirs(output_dir, exist_ok=True)
    filename = f"shift_{year}_{month:02d}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath
